import openpyxl
import copy
from operator import itemgetter, attrgetter


class Read:
    name = 'aaa'

    def cmmn_std_excel(self, file):
        workbook = openpyxl.load_workbook(file)
        sns = workbook.sheetnames
        ws = workbook.worksheets[1]
        domain = self.cmmn_std_domain_excel(ws)

        ws = workbook.worksheets[0]
        term = self.cmmn_std_term_excel(ws, domain)

        return term

    def cmmn_std_domain_excel(self, ws):

        domain = {}

        row_idx = 0

        for row in ws:
            if row_idx == 0:
                row_idx += 1
                continue

            domain_data = {}

            cell_idx = 0

            for cell in row:
                if cell_idx == 3:
                    domain_data['domNm'] = self.get_value(cell.value)
                elif cell_idx == 5:
                    domain_data['datTp'] = self.get_value(cell.value)
                elif cell_idx == 6:
                    if not cell.value or cell.value == '0' or cell.value == '-':
                        domain_data['datLen'] = ''
                    else:
                        domain_data['datLen'] = self.get_value(cell.value)
                elif cell_idx == 7:
                    if not cell.value or cell.value == '0' or cell.value == '-':
                        domain_data['datDcmlLen'] = ''
                    else:
                        domain_data['datDcmlLen'] = self.get_value(cell.value)

                cell_idx += 1

            domain[domain_data['domNm']] = domain_data

            row_idx += 1
        return domain

    def cmmn_std_term_excel(self, ws, domain):
        term = []
        row_idx = 0

        for row in ws:
            if row_idx == 0:
                row_idx += 1
                continue

            term_data = {}

            cell_idx = 0

            for cell in row:
                if cell_idx == 1:
                    term_data['termNm'] = self.get_value(cell.value)
                elif cell_idx == 2:
                    term_data['termEngNm'] = self.get_value(cell.value)
                elif cell_idx == 3:
                    if not cell.value:
                        raise Exception('용어에 도메인항목이 비어있습니다.')

                    dom_nm = cell.value
                    term_data['domNm'] = dom_nm
                    if (dom_nm not in domain):
                        raise Exception('해당도메인이 존재하지 않습니다.')
                    term_data['datTp'] = domain[dom_nm]['datTp']
                    term_data['datLen'] = domain[dom_nm]['datLen']
                    term_data['datDcmlLen'] = domain[dom_nm]['datDcmlLen']
                elif cell_idx == 4:
                    term_data['termDesc'] = self.get_value(cell.value)

                cell_idx += 1

            term.append(term_data)
            row_idx += 1
        return term

    def std_excel(self, file):
        workbook = openpyxl.load_workbook(file)
        sns = workbook.sheetnames
        ws = workbook.worksheets[0]

        term = []
        row_idx = 0

        for row in ws:
            if row_idx == 0:
                row_idx += 1
                continue

            term_data = {}

            cell_idx = 0

            for cell in row:
                if cell_idx == 1:
                    term_data['termNm'] = self.get_value(cell.value)
                elif cell_idx == 2:
                    term_data['termEngNm'] = self.get_value(cell.value)
                elif cell_idx == 3:
                    term_data['datTp'] = self.get_value(cell.value)
                elif cell_idx == 4:
                    if not cell.value or str(cell.value) == '0' or cell.value == '-':
                        term_data['datLen'] = ''
                    else:
                        term_data['datLen'] = self.get_value(cell.value)
                elif cell_idx == 5:
                    if not cell.value or str(cell.value) == '0' or cell.value == '-':
                        term_data['datDcmlLen'] = ''
                    else:
                        term_data['datDcmlLen'] = self.get_value(cell.value)
                elif cell_idx == 6:
                    term_data['termDesc'] = self.get_value(cell.value)
                elif cell_idx == 7:
                    term_data['domNm'] = self.get_value(cell.value)

                cell_idx += 1

            term.append(term_data)

            row_idx += 1
        return term

    def compare_excel(self, std, cmmn_std):

        cmmn_std_map = {}
        exist_term_nm = {}
        compare_term = []

        for term in cmmn_std:
            cmmn_std_map[term['termNm']] = term

        for term in std:

            new_term = copy.deepcopy(term)
            term_nm = term['termNm']
            if (term_nm in cmmn_std_map):
                cmmn_std_term = cmmn_std_map[term_nm]

                new_term['cmmnTermNm'] = cmmn_std_term['termNm']
                new_term['cmmnTermEngNm'] = cmmn_std_term['termEngNm']
                new_term['cmmnDatTp'] = cmmn_std_term['datTp']
                new_term['cmmnDatLen'] = cmmn_std_term['datLen']
                new_term['cmmnDatDcmlLen'] = cmmn_std_term['datDcmlLen']
                new_term['cmmnDomNm'] = cmmn_std_term['domNm']

                if term['termEngNm'] == cmmn_std_term['termEngNm'] and term['datTp'] == cmmn_std_term['datTp'] and term['datLen'] == cmmn_std_term['datLen'] and term['datDcmlLen'] == cmmn_std_term['datDcmlLen']:
                    new_term['stdYn'] = 'Y'
                    new_term['desc'] = ''
                else:
                    new_term['stdYn'] = 'N'
                    if term['termEngNm'] != cmmn_std_term['termEngNm']:
                        new_term['desc'] = '영문약어불일치'
                    elif term['datTp'] != cmmn_std_term['datTp']:
                        new_term['desc'] = '데이터타입불일치'
                    elif term['datLen'] != cmmn_std_term['datLen'] or term['datDcmlLen'] != cmmn_std_term['datDcmlLen']:
                        new_term['desc'] = '데이터길이불일치'

                exist_term_nm[term_nm] = 'O'
            else:
                new_term['cmmnTermNm'] = ''
                new_term['cmmnTermEngNm'] = ''
                new_term['cmmnDatTp'] = ''
                new_term['cmmnDatLen'] = ''
                new_term['cmmnDatDcmlLen'] = ''
                new_term['cmmnDomNm'] = ''
                new_term['stdYn'] = ''
                new_term['desc'] = ''

            compare_term.append(new_term)

        for term in cmmn_std:
            term_nm = term['termNm']
            if (term_nm not in exist_term_nm):
                new_term = {}
                new_term['termNm'] = ''
                new_term['termEngNm'] = ''
                new_term['datTp'] = ''
                new_term['datLen'] = ''
                new_term['datDcmlLen'] = ''
                new_term['domNm'] = ''
                new_term['termDesc'] = ''
                new_term['cmmnTermNm'] = term['termNm']
                new_term['cmmnTermEngNm'] = term['termEngNm']
                new_term['cmmnDatTp'] = term['datTp']
                new_term['cmmnDatLen'] = term['datLen']
                new_term['cmmnDatDcmlLen'] = term['datDcmlLen']
                new_term['cmmnDomNm']= term['domNm']
                new_term['stdYn'] = 'N'
                new_term['desc'] = ''

                compare_term.append(new_term);

        return compare_term;

    def merge_excel(self, std, cmmn_std):

        cmmn_std_map = {}
        same_term_nm = {}
        term_list = []

        for term in cmmn_std:
            cmmn_std_map[term['termNm']] = term

        for term in std:

            new_term = copy.deepcopy(term)

            term_nm = term['termNm']

            if term_nm in cmmn_std_map:

                cmmn_std_term = cmmn_std_map[term_nm]

                if term['termEngNm'] == cmmn_std_term['termEngNm'] and term['datTp'] == cmmn_std_term['datTp'] and term['datLen'] == cmmn_std_term['datLen'] and term['datDcmlLen'] == cmmn_std_term['datDcmlLen']:

                    new_term['mappingTermNm'] = ''
                    new_term['mappingTermEngNm'] = ''
                    new_term['mappingDatTp'] = ''
                    new_term['mappingDatLen'] = ''
                    new_term['mappingDatDcmlLen'] = ''
                    new_term['mappingDomNm'] = ''

                    new_term['stdYn'] = 'Y'
                    new_term['dv'] = '공통표준'
                    new_term['desc'] = ''
                    new_term['ord'] = new_term['termNm']

                    same_term_nm[term_nm] = 'O'

                else:

                    new_term['mappingTermNm'] = cmmn_std_term['termNm']
                    new_term['mappingTermEngNm'] = cmmn_std_term['termEngNm']
                    new_term['mappingDatTp'] = cmmn_std_term['datTp']
                    new_term['mappingDatLen'] = cmmn_std_term['datLen']
                    new_term['mappingDatDcmlLen'] = cmmn_std_term['datDcmlLen']
                    new_term['mappingDomNm'] = cmmn_std_term['domNm']

                    new_term['stdYn'] = 'N'
                    new_term['dv'] = '기관표준'
                    new_term['ord'] = new_term['termNm']+'2'

                    if term['termEngNm'] != cmmn_std_term['termEngNm']:
                        new_term['desc'] = '영문약어불일치'
                    elif term['datTp'] != cmmn_std_term['datTp']:
                        new_term['desc'] = '데이터타입불일치'
                    elif term['datLen'] != cmmn_std_term['datLen'] or term['datDcmlLen'] != cmmn_std_term['datDcmlLen']:
                        new_term['desc'] = '데이터길이불일치'

            else:

                new_term['mappingTermNm'] = ''
                new_term['mappingTermEngNm'] = ''
                new_term['mappingDatTp'] = ''
                new_term['mappingDatLen'] = ''
                new_term['mappingDatDcmlLen'] = ''
                new_term['mappingDomNm'] = ''
                new_term['stdYn'] = 'Y'
                new_term['dv'] = '기관표준'
                new_term['desc'] = ''
                new_term['ord'] = new_term['termNm']

            term_list.append(new_term)

        for term in cmmn_std:
            term_nm = term['termNm']
            if term_nm not in same_term_nm:
                new_term = {}
                new_term['termNm'] = term['termNm']
                new_term['termEngNm'] = term['termEngNm']
                new_term['datTp'] = term['datTp']
                new_term['datLen'] = term['datLen']
                new_term['datDcmlLen'] = term['datDcmlLen']
                new_term['domNm'] = term['domNm']

                new_term['mappingTermNm'] = ''
                new_term['mappingTermEngNm'] = ''
                new_term['mappingDatTp'] = ''
                new_term['mappingDatLen'] = ''
                new_term['mappingDatDcmlLen'] = ''
                new_term['mappingDomNm'] = ''
                new_term['stdYn'] = 'Y'
                new_term['dv'] = '공통표준'
                new_term['desc'] = ''
                new_term['ord'] = new_term['termNm']+'1'

                term_list.append(new_term);

        return sorted(term_list, key=itemgetter('ord'))

    def get_value(self, val):
        if val:
            return str(val)
        else:
            return ''
