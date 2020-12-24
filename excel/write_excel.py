import datetime
import openpyxl

class Write:  # 클래스

    def compare_excel(self, data, file):

        wb = openpyxl.Workbook()
        for sheet in wb.sheetnames:
            wb.remove(wb[sheet])

        ws = wb.create_sheet(title='데이터표준비교', index=0)

        ws.merge_cells('A1:A2')
        ws.merge_cells('B1:F1')
        ws.merge_cells('G1:K1')
        ws.merge_cells('L1:L2')
        ws.merge_cells('M1:M2')

        ws['A1'] = '번호'
        ws['B1'] = '기관표준용어'
        ws['G1'] = '공통표준용어'
        ws['L1'] = '표준적용여부'
        ws['M1'] = '비고'

        ws['B2'] = '기관표준용어명'
        ws['C2'] = '영문약어명'
        ws['D2'] = '데이터유형'
        ws['E2'] = '데이터길이'
        ws['F2'] = '소수점데이터길이'
        ws['G2'] = '공통표준용어명'
        ws['H2'] = '영문약어명'
        ws['I2'] = '데이터유형'
        ws['J2'] = '데이터길이'
        ws['K2'] = '소수점데이터길이'

        for row in ws:
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.fills.PatternFill(patternType='solid',
                                                              fgColor=openpyxl.styles.colors.Color(rgb='CCCCCC'))

        for col in range(ws.max_column):
            ws.column_dimensions[chr(ord('A') + col)].width = 15

        start_row = 3
        for (row, item) in enumerate(data, start_row):
            ws.cell(row=row, column=1, value=row - 2)
            ws.cell(row=row, column=2, value=self.getData(item, 'termNm'))
            ws.cell(row=row, column=2, value=self.getData(item, 'termNm'))
            ws.cell(row=row, column=3, value=self.getData(item, 'termEngNm'))
            ws.cell(row=row, column=4, value=self.getData(item, 'datTp'))
            ws.cell(row=row, column=5, value=self.getData(item, 'datLen'))
            ws.cell(row=row, column=6, value=self.getData(item, 'datDcmlLen'))
            ws.cell(row=row, column=7, value=self.getData(item, 'cmmnTermNm'))
            ws.cell(row=row, column=8, value=self.getData(item, 'cmmnTermEngNm'))
            ws.cell(row=row, column=9, value=self.getData(item, 'cmmnDatTp'))
            ws.cell(row=row, column=10, value=self.getData(item, 'cmmnDatLen'))
            ws.cell(row=row, column=11, value=self.getData(item, 'cmmnDatDcmlLen'))
            ws.cell(row=row, column=12, value=self.getData(item, 'stdYn'))
            ws.cell(row=row, column=13, value=self.getData(item, 'desc'))

        for row in ws:
            for cell in row:
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style="thin", color='FF000000'),
                                     right=openpyxl.styles.Side(border_style="thin", color='FF000000'),
                                     top=openpyxl.styles.Side(border_style="thin", color='FF000000'),
                                     bottom=openpyxl.styles.Side(border_style="thin", color='FF000000'))
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        wb.save(file)
        wb.close()

    def merge_excel(self, data, file):
        wb = openpyxl.Workbook()
        for sheet in wb.sheetnames:
            wb.remove(wb[sheet])

        ws = wb.create_sheet(title='데이터표준비교', index=0)

        ws.merge_cells('A1:A2')
        ws.merge_cells('B1:G1')
        ws.merge_cells('H1:H2')
        ws.merge_cells('I1:M1')
        ws.merge_cells('N1:N2')

        ws['A1'] = '번호'
        ws['B1'] = '표준용어'
        ws['I1'] = '비표준용어매핑'
        ws['N1'] = '비고'

        ws['B2'] = '구분'
        ws['C2'] = '기관표준용어명'
        ws['D2'] = '영문약어명'
        ws['E2'] = '데이터유형'
        ws['F2'] = '데이터길이'
        ws['G2'] = '소수점데이터길이'

        ws['H1'] = '표준여부'

        ws['I2'] = '표준용어명'
        ws['J2'] = '영문약어명'
        ws['K2'] = '데이터유형'
        ws['L2'] = '데이터길이'
        ws['M2'] = '소수점데이터길이'

        for row in ws:
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.fills.PatternFill(patternType='solid',
                                                              fgColor=openpyxl.styles.colors.Color(rgb='CCCCCC'))

        for col in range(ws.max_column):
            ws.column_dimensions[chr(ord('A') + col)].width = 15

        start_row = 3
        for (row, item) in enumerate(data, start_row):
            ws.cell(row=row, column=1, value=row - 2)
            ws.cell(row=row, column=2, value=self.getData(item, 'dv'))
            ws.cell(row=row, column=3, value=self.getData(item, 'termNm'))
            ws.cell(row=row, column=4, value=self.getData(item, 'termEngNm'))
            ws.cell(row=row, column=5, value=self.getData(item, 'datTp'))
            ws.cell(row=row, column=6, value=self.getData(item, 'datLen'))
            ws.cell(row=row, column=7, value=self.getData(item, 'datDcmlLen'))
            ws.cell(row=row, column=8, value=self.getData(item, 'stdYn'))
            ws.cell(row=row, column=9, value=self.getData(item, 'mappingTermNm'))
            ws.cell(row=row, column=10, value=self.getData(item, 'mappingTermEngNm'))
            ws.cell(row=row, column=11, value=self.getData(item, 'mappingDatTp'))
            ws.cell(row=row, column=12, value=self.getData(item, 'mappingDatLen'))
            ws.cell(row=row, column=13, value=self.getData(item, 'mappingDatDcmlLen'))
            ws.cell(row=row, column=14, value=self.getData(item, 'desc'))

        for row in ws:
            for cell in row:
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style="thin", color='FF000000'),
                                     right=openpyxl.styles.Side(border_style="thin", color='FF000000'),
                                     top=openpyxl.styles.Side(border_style="thin", color='FF000000'),
                                     bottom=openpyxl.styles.Side(border_style="thin", color='FF000000'))
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        wb.save(file)
        wb.close()

    def getData(self, item, keyStr):
        if keyStr in item:
            return item[keyStr]
        else:
            return ''






